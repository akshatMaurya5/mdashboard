from __future__ import annotations

import json
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE_PRIORITY = ["2026", "backup", "history"]


@dataclass(frozen=True)
class DailyValue:
    j: int
    a: int
    m: int

    @property
    def total(self) -> int:
        return self.j + self.a + self.m


def to_count(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(round(float(value)))
    except Exception:
        return 0


def to_iso_date(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return None


def parse_history(ws) -> dict[str, DailyValue]:
    out: dict[str, DailyValue] = {}
    for r in range(3, ws.max_row + 1):
        date_key = to_iso_date(ws.cell(r, 1).value)
        if not date_key:
            continue
        out[date_key] = DailyValue(
            j=to_count(ws.cell(r, 5).value),
            a=to_count(ws.cell(r, 6).value),
            m=to_count(ws.cell(r, 7).value),
        )
    return out


def parse_column_blocks(ws, starts: list[int], row_start: int = 3) -> dict[str, DailyValue]:
    out: dict[str, DailyValue] = {}
    for c in starts:
        for r in range(row_start, ws.max_row + 1):
            date_key = to_iso_date(ws.cell(r, c).value)
            if not date_key:
                continue
            out[date_key] = DailyValue(
                j=to_count(ws.cell(r, c + 1).value),
                a=to_count(ws.cell(r, c + 2).value),
                m=to_count(ws.cell(r, c + 3).value),
            )
    return out


def source_summary(rows: dict[str, DailyValue]) -> dict:
    if not rows:
        return {"days": 0, "from": None, "to": None}
    keys = sorted(rows.keys())
    return {"days": len(rows), "from": keys[0], "to": keys[-1]}


def build_dataset(workbook_path: Path) -> dict:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    parsed_sources: dict[str, dict[str, DailyValue]] = {
        "history": parse_history(wb["history"]),
        "backup": parse_column_blocks(wb["backup"], [1, 5, 9, 13], row_start=3),
        "2026": parse_column_blocks(wb["2026"], [1, 5, 9], row_start=3),
    }

    all_dates = sorted(set().union(*[set(s.keys()) for s in parsed_sources.values()]))
    chosen_rows = []
    conflict_dates = []
    duplicate_date_count = 0

    for date_key in all_dates:
        available: dict[str, DailyValue] = {
            source_name: rows[date_key]
            for source_name, rows in parsed_sources.items()
            if date_key in rows
        }
        if len(available) > 1:
            duplicate_date_count += 1

        distinct_values = {(v.j, v.a, v.m) for v in available.values()}
        has_conflict = len(distinct_values) > 1
        if has_conflict:
            conflict_dates.append(
                {
                    "date": date_key,
                    "valuesBySource": {
                        k: {"j": v.j, "a": v.a, "m": v.m}
                        for k, v in sorted(available.items())
                    },
                }
            )

        selected_source = None
        selected_value = None
        for source_name in SOURCE_PRIORITY:
            if source_name in available:
                selected_source = source_name
                selected_value = available[source_name]
                break

        if selected_source is None or selected_value is None:
            continue

        chosen_rows.append(
            {
                "date": date_key,
                "j": selected_value.j,
                "a": selected_value.a,
                "m": selected_value.m,
                "total": selected_value.total,
                "source": selected_source,
                "sourceCount": len(available),
                "conflict": has_conflict,
            }
        )

    return {
        "generatedAt": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "workbookPath": str(workbook_path),
        "sourcePriority": SOURCE_PRIORITY,
        "records": chosen_rows,
        "diagnostics": {
            "totalDays": len(chosen_rows),
            "dateRange": {
                "from": chosen_rows[0]["date"] if chosen_rows else None,
                "to": chosen_rows[-1]["date"] if chosen_rows else None,
            },
            "duplicatedDateCount": duplicate_date_count,
            "conflictCount": len(conflict_dates),
            "conflicts": conflict_dates,
            "sourceCoverage": {
                source_name: source_summary(rows)
                for source_name, rows in parsed_sources.items()
            },
        },
    }


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python scripts/build_dataset.py <workbook_path> [output_json_path]")
        return 1

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = (
        Path(sys.argv[2]).expanduser().resolve()
        if len(sys.argv) > 2
        else Path("src/data/coomersData.json").resolve()
    )

    dataset = build_dataset(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(dataset, indent=2), encoding="utf-8")

    print(f"Wrote {len(dataset['records'])} records to {output_path}")
    print(f"Duplicates: {dataset['diagnostics']['duplicatedDateCount']}")
    print(f"Conflicts: {dataset['diagnostics']['conflictCount']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
